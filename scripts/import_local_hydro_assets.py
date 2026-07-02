#!/usr/bin/env python3
"""Import local Puerto Rico hydro datasets into AguaYLuz utility_assets.jsonl.

Supported sources:
  * PRI.gpkg layers: water_treatment_plant, wastewater_plant, pumping_station,
    pipeline, pipeline_feature
  * Acueductos & Canales de Riego CSVs: PRASA_Intakes_Outfalls_v1.csv,
    Waterworks_Integrated_v2.csv, CANAL_SYSTEMS.csv, CANAL_DE_RIEGO_SEGMENTS.csv,
    Conduit_Alignments_v0.csv
  * NID_AUTH_MASTER.geojson / any NID GeoJSON

The importer is deterministic and idempotent: rows are keyed by asset_id and merge
into data/utility_assets.jsonl without duplicating prior live EPA/USGS/EIA/HIFLD rows.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import shutil
import sqlite3
import sys
import tempfile
import zipfile
import tarfile
from collections import Counter
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "data"
OUT = DATA / "utility_assets.jsonl"
PR_BOUNDS = (-67.95, 17.7, -65.2, 18.7)  # lon min, lat min, lon max, lat max
PRI_LAYERS = {
    "water_treatment_plant": ("water", "treatment_plant", "point"),
    "wastewater_plant": ("wastewater", "wastewater_plant", "point"),
    "pumping_station": ("water", "pump_station", "point"),
    "pipeline": ("water", "pipeline", "line"),
    "pipeline_feature": ("water", "pipeline_feature", "unknown"),
    "water_tower_point": ("water", "water_tower", "point"),
    "water_tower_polygon": ("water", "water_tower", "polygon"),
    "water_well_point": ("water", "water_well", "point"),
    "water_well_polygon": ("water", "water_well", "polygon"),
    "water_reservoir": ("water", "reservoir", "polygon"),
}
# Coastal outlet GPKG layers — QA-confirmed outlet points from PR_COASTAL_OUTLET_MASTER bundle
# (natural-hydrography reference layers like aquifer boundaries are out of scope: aguayluz-pr
# ingests water measurements/infrastructure, not general hydrography mapping)
COASTAL_OUTLET_LAYERS = {
    "qa_outlet_points_v2_1_all":       ("water", "coastal_outlet", "point", "AguaYLuz QA"),
    "promoted_confirmed_points_v2_1":  ("water", "coastal_outlet_confirmed", "point", "AguaYLuz QA"),
    "southwest_aoi_outlets":           ("water", "coastal_outlet", "point", "AguaYLuz QA"),
    "gap_queue_points":                ("water", "coastal_outlet_candidate", "point", "AguaYLuz QA"),
}
CSV_SPECS = {
    "PRASA_Intakes_Outfalls_v1.csv": ("water", "intake_outfall", "point", "PRASA"),
    "Waterworks_Integrated_v2.csv": ("water", "waterworks", "unknown", "AAA/PRASA"),
    "CANAL_SYSTEMS.csv": ("water", "canal_system", "line", "irrigation"),
    "CANAL_DE_RIEGO_SEGMENTS.csv": ("water", "irrigation_canal", "line", "irrigation"),
    "Conduit_Alignments_v0.csv": ("water", "conduit_alignment", "line", "unknown"),
    "PreAAA_Aqueducts_Master_Register.csv": ("water", "historic_aqueduct", "line", "Pre-AAA"),
    "PreAAA_Aqueducts_Master_Register_v2.csv": ("water", "historic_aqueduct", "line", "Pre-AAA"),
    "PreAAA_Aqueducts_Master_Register_v2_1.csv": ("water", "historic_aqueduct", "line", "Pre-AAA"),
    "Canal_de_Riego_features_summary.csv": ("water", "canal_feature", "unknown", "irrigation"),
    # Coastal outlet confirmed points (PR_COASTAL_OUTLET_MASTER_v2_outputs.zip)
    "PR_COASTAL_OUTLET_MASTER_v2_1_confirmed_points.csv": ("water", "coastal_outlet_confirmed", "point", "AguaYLuz QA"),
    # EPA contamination sites with UGCN corridor cross-reference (PR_Environmental_Protection.zip)
    "PR_Contamination_Master_Full625_SEMSmerged.csv": ("water", "contamination_site", "point", "EPA SEMS"),
    "PR_EPA_SEMS_PR_ACTU_502.csv": ("water", "sems_site", "point", "EPA SEMS"),
    # Dropped (out of scope): karst/subsurface features (FEATURE_MASTER.csv, KARST_EXTENSION.csv),
    # UGCN/WIDL/ICG conduit-corridor and vulnerability outputs (UGCN_MASTER_ILAP_TABLE_v1.csv,
    # widl_nodes.csv, icg_priority_zones.csv, Hydro_ILAP_Node_Vulnerability_Table.csv) — these are
    # subsurface-corridor inference products, which federation_manifest.yaml assigns to
    # spiderweb-pr, not aguayluz-pr's water measurement/infrastructure scope.
}

NID_LAYERS = {"Dams"}
# NHDPlus hydrography (flowlines/catchments/areas), USFWS wetlands, and USGS GNIS
# gazetteer place-names are natural-hydrography reference/geometry layers, not water
# infrastructure or measurement sites — out of aguayluz-pr's scope. Dropped on purpose.

MAX_IMPORT_ROWS_BY_LAYER: dict[str, int] = {}
SHAPEFILE_SPECS = {
    # Lago Carite reservoir shoreline (Carite_shr83.zip) — a named reservoir's shoreline,
    # unlike the dropped NHDPlus hydro-enforcement processing artifacts (Burn*/Sink/Wall/LandSea).
    "Carite_shr83.shp": ("water", "reservoir_shoreline", "polygon", "USGS/PRASA"),
}
XLSX_SHEET_SPECS = {
    # RESERVOIRS, WWTP/wastewater infra, NPDES facilities/outfalls, springs, gage/monitoring
    # sites, dams, and planned reservoirs are infrastructure or direct water-measurement sites.
    # HYDRO_FEATURES_FULL/EPA_NPDES_CATCHMENTS_PR/SURFACE_LINEARWATER/HL_WATERSHED_CONTEXT are
    # generic hydrography reference buckets — dropped, out of scope.
    "RESERVOIRS": ("water", "reservoir", "point", "PR Hydro Repository"),
    "WWTP_CORE_FULL": ("wastewater", "wastewater_plant", "point", "PR Hydro Repository"),
    "WASTEWATER_INFRA_FULL": ("wastewater", "wastewater_asset", "unknown", "PR Hydro Repository"),
    "EPA_NPDES_FACILITIES_PR": ("wastewater", "npdes_facility", "point", "EPA NPDES"),
    "EPA_NPDES_OUTFALLS_PR": ("wastewater", "npdes_outfall", "point", "EPA NPDES"),
    "SPRING_IDENTITY_CROSSWALK": ("water", "spring", "point", "PR Hydro Repository"),
    "USGS_GW_LEVEL_SITES": ("water", "groundwater_monitoring_site", "point", "USGS"),
    "HL_USGS_NWIS_KeyGages_PR": ("water", "usgs_key_gage", "point", "USGS"),
    "HL_PR_DamNodes_EXPANDED": ("water", "dam", "point", "PR Hydro Repository"),
    "HL_PLANNED_RESERVOIRS": ("water", "planned_reservoir", "point", "PR Hydro Repository"),
    "HL_ILAP_NODES": ("water", "ilap_hydro_node", "point", "PR Hydro Repository"),
    "EIA_PR_Hydro_Inventory": ("power", "hydroelectric_generator", "point", "EIA"),
}


def stable_id(prefix: str, *parts: Any) -> str:
    h = hashlib.sha256("|".join(str(p) for p in parts if p is not None).encode()).hexdigest()[:16]
    return f"{prefix}_{h}"


@lru_cache(maxsize=256)
def source_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in {"nan", "none", "null", "<null>"}:
        return None
    return s


def as_float(v: Any) -> float | None:
    try:
        if v in (None, ""):
            return None
        x = float(v)
        if math.isnan(x):
            return None
        return x
    except (TypeError, ValueError):
        return None


def in_pr(lat: float | None, lon: float | None) -> bool:
    if lat is None or lon is None:
        return False
    return PR_BOUNDS[1] <= lat <= PR_BOUNDS[3] and PR_BOUNDS[0] <= lon <= PR_BOUNDS[2]


def pick(row: dict[str, Any], keys: Iterable[str]) -> Any:
    lower = {str(k).lower(): k for k in row}
    for key in keys:
        k = lower.get(key.lower())
        if k is not None and clean(row.get(k)) is not None:
            return row.get(k)
    return None


def latlon_from_row(row: dict[str, Any]) -> tuple[float | None, float | None]:
    lat = as_float(pick(row, ["lat", "latitude", "y", "LATITUDE", "POINT_Y", "GEOCODE_LATITUDE", "outfall_lat", "Lat", "lat_dd", "prim_lat_dec"]))
    lon = as_float(pick(row, ["lon", "long", "longitude", "x", "LONGITUDE", "POINT_X", "GEOCODE_LONGITUDE", "outfall_lon", "Lon", "lon_dd", "prim_long_dec"]))
    if in_pr(lat, lon):
        return lat, lon
    # Common reversed coordinate accident.
    if in_pr(lon, lat):
        return lon, lat
    return None, None


def centroid_from_geojson_geometry(geom: dict[str, Any] | None) -> tuple[float | None, float | None, str]:
    if not geom:
        return None, None, "unknown"
    typ = geom.get("type") or "unknown"
    coords = geom.get("coordinates")
    pts: list[tuple[float, float]] = []

    def walk(obj: Any) -> None:
        if isinstance(obj, (list, tuple)) and len(obj) >= 2 and all(isinstance(x, (int, float)) for x in obj[:2]):
            pts.append((float(obj[0]), float(obj[1])))
        elif isinstance(obj, (list, tuple)):
            for x in obj:
                walk(x)
    walk(coords)
    if not pts:
        return None, None, typ.lower()
    lon = sum(p[0] for p in pts) / len(pts)
    lat = sum(p[1] for p in pts) / len(pts)
    gtype = "point" if "Point" in typ else "line" if "Line" in typ else "polygon" if "Polygon" in typ else "unknown"
    return (round(lat, 6), round(lon, 6), gtype) if in_pr(lat, lon) else (None, None, gtype)


def make_asset(*, prefix: str, source_path: Path, source_ref: str, row: dict[str, Any], asset_type: str,
               subtype: str, geometry_type: str, operator: str | None = None, lat: float | None = None,
               lon: float | None = None, confidence: int = 85, evidence_tier: str = "T1") -> dict[str, Any]:
    name = clean(pick(row, ["name", "asset_name", "facility_name", "facility_name_norm", "FACILITY", "Facility", "NAME", "feature_name", "dam_name", "NIDID", "id", "ID", "Station_Name", "Plant Name", "Reservoir_Name", "NodeName"]))
    muni = clean(pick(row, ["municipality", "municipio", "city", "CITY", "county", "COUNTY", "county_name", "MUNICIPIO", "County", "Basin"])) or "Puerto Rico"
    op = clean(operator) or clean(pick(row, ["operator", "owner", "OWNER", "OWNER_NAME", "owner_operator", "PRIMARY_OWNER_TYPE", "agency", "utility"]))
    raw_key = clean(pick(row, ["source_row_number", "asset_id", "asset_uid", "facility_uid", "wwtp_uid", "outfall_uid", "Reservoir_ID", "Station_ID", "Plant ID", "pr_id", "feature_id", "id", "ID", "OBJECTID", "fid", "NIDID", "nidid", "nhdplusid", "source_id"]))
    if raw_key is None:
        raw_key = stable_id("ROW", source_ref, json.dumps(row, sort_keys=True, default=str)[:1000])
    asset_id = stable_id(prefix, source_ref, raw_key)
    return {
        "asset_id": asset_id,
        "asset_name": name or f"{subtype} {raw_key}",
        "asset_type": asset_type,
        "asset_subtype": subtype,
        "operator": op,
        "municipality": muni,
        "lat": lat,
        "lon": lon,
        "geometry_type": geometry_type,
        "status": "unknown",
        "source_ref": source_ref,
        "source_hash": source_hash(source_path),
        "evidence_tier": evidence_tier,
        "confidence": confidence,
        "review_status": "accepted" if confidence >= 80 else "needs_review",
        "attribute_coverage": "partial",
    }


def iter_geojson(path: Path, prefix: str, subtype: str, asset_type: str = "water") -> Iterable[dict[str, Any]]:
    doc = json.loads(path.read_text(encoding="utf-8"))
    for feat in doc.get("features", []):
        props = feat.get("properties") or {}
        lat, lon, gtype = centroid_from_geojson_geometry(feat.get("geometry"))
        yield make_asset(prefix=prefix, source_path=path, source_ref=path.name, row=props, asset_type=asset_type,
                         subtype=subtype, geometry_type=gtype, operator=pick(props, ["owner", "OWNER_NAME"]),
                         lat=lat, lon=lon, confidence=90)


def iter_csv(path: Path, spec: tuple[str, str, str, str]) -> Iterable[dict[str, Any]]:
    asset_type, subtype, gtype, operator = spec
    with path.open(newline="", encoding="utf-8-sig") as f:
        for source_row_number, row in enumerate(csv.DictReader(f), start=1):
            row["source_row_number"] = source_row_number
            lat, lon = latlon_from_row(row)
            yield make_asset(prefix="LOCAL", source_path=path, source_ref=path.name, row=row, asset_type=asset_type,
                             subtype=subtype, geometry_type=gtype, operator=operator, lat=lat, lon=lon,
                             confidence=82 if lat is not None else 72, evidence_tier="T2")


def gpkg_layers(path: Path) -> list[str]:
    with sqlite3.connect(path) as con:
        return [r[0] for r in con.execute("select table_name from gpkg_contents where data_type='features'")]


def _layer_spec(path: Path, layer: str) -> tuple[str, str, str, str] | None:
    if layer in PRI_LAYERS:
        asset_type, subtype, gtype = PRI_LAYERS[layer]
        return asset_type, subtype, gtype, "PRASA/AAA"
    if layer in NID_LAYERS:
        return "water", "dam", "point", "National Inventory of Dams"
    if layer in COASTAL_OUTLET_LAYERS:
        return COASTAL_OUTLET_LAYERS[layer]
    return None


def _row_from_geoseries(rec: Any) -> dict[str, Any]:
    return {k: v for k, v in rec.items() if k != "geometry"}


def _geom_latlon_type(geom: Any, default_gtype: str) -> tuple[float | None, float | None, str]:
    lat = lon = None
    gtype = default_gtype
    if geom is not None and not geom.is_empty:
        pt = geom.representative_point()
        lat, lon = round(float(pt.y), 6), round(float(pt.x), 6)
        geom_name = geom.geom_type.lower()
        if geom_name.endswith("point") or "point" in geom_name:
            gtype = "point"
        elif "line" in geom_name:
            gtype = "line"
        elif "polygon" in geom_name:
            gtype = "polygon"
    return (lat if in_pr(lat, lon) else None, lon if in_pr(lat, lon) else None, gtype)


def iter_gpkg(path: Path, layers: list[str] | None = None) -> Iterable[dict[str, Any]]:
    try:
        import geopandas as gpd  # type: ignore
    except Exception as exc:  # pragma: no cover - environment guard
        raise RuntimeError("geopandas is required for GPKG ingestion") from exc
    selected = layers or gpkg_layers(path)
    for layer in selected:
        spec = _layer_spec(path, layer)
        if spec is None:
            continue
        asset_type, subtype, default_gtype, operator = spec
        gdf = gpd.read_file(path, layer=layer).to_crs(4326)
        max_rows = MAX_IMPORT_ROWS_BY_LAYER.get(layer)
        for n, (idx, rec) in enumerate(gdf.iterrows()):
            if max_rows is not None and n >= max_rows:
                break
            row = _row_from_geoseries(rec)
            lat, lon, gtype = _geom_latlon_type(rec.geometry, default_gtype)
            row.setdefault("fid", idx)
            confidence = 92 if lat is not None else 76
            yield make_asset(prefix="LOCAL", source_path=path, source_ref=f"{path.name}:{layer}", row=row,
                             asset_type=asset_type, subtype=subtype, geometry_type=gtype, operator=operator,
                             lat=lat, lon=lon, confidence=confidence, evidence_tier="T1")


def iter_shapefile(path: Path, spec: tuple[str, str, str, str]) -> Iterable[dict[str, Any]]:
    try:
        import geopandas as gpd  # type: ignore
    except Exception as exc:  # pragma: no cover - environment guard
        raise RuntimeError("geopandas is required for shapefile ingestion") from exc
    asset_type, subtype, default_gtype, operator = spec
    gdf = gpd.read_file(path).to_crs(4326)
    for idx, rec in gdf.iterrows():
        row = _row_from_geoseries(rec)
        lat, lon, gtype = _geom_latlon_type(rec.geometry, default_gtype)
        row.setdefault("fid", idx)
        yield make_asset(prefix="LOCAL", source_path=path, source_ref=path.name, row=row,
                         asset_type=asset_type, subtype=subtype, geometry_type=gtype, operator=operator,
                         lat=lat, lon=lon, confidence=88 if lat is not None else 72, evidence_tier="T1")


def _sheet_header_and_rows(ws: Any) -> tuple[list[str], list[tuple[Any, ...]]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows[:12]):
        values = [clean(v) for v in row]
        named = [v for v in values if v]
        has_coord = any(str(v).lower() in {"lat", "latitude", "geocode_latitude", "outfall_lat", "lat_dd", "prim_lat_dec"} for v in named)
        has_name = any("name" in str(v).lower() or str(v).lower() in {"facility_id", "asset_uid", "nidid", "station_id"} for v in named)
        if len(named) >= 3 and (has_coord or has_name):
            header_idx = i
            break
    if header_idx is None:
        return [], []
    headers = [clean(v) or f"col_{j}" for j, v in enumerate(rows[header_idx])]
    return headers, rows[header_idx + 1:]


def iter_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl is required for XLSX ingestion") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet, spec in XLSX_SHEET_SPECS.items():
        if sheet not in wb.sheetnames:
            continue
        asset_type, subtype, gtype, operator = spec
        ws = wb[sheet]
        headers, rows = _sheet_header_and_rows(ws)
        if not headers:
            continue
        for source_row_number, values in enumerate(rows, start=1):
            row = dict(zip(headers, values))
            if not any(clean(v) for v in row.values()):
                continue
            row["source_row_number"] = source_row_number
            row["source_sheet"] = sheet
            lat, lon = latlon_from_row(row)
            confidence = 88 if lat is not None else 70
            yield make_asset(prefix="XLSX", source_path=path, source_ref=f"{path.name}:{sheet}", row=row,
                             asset_type=asset_type, subtype=subtype, geometry_type=gtype, operator=operator,
                             lat=lat, lon=lon, confidence=confidence, evidence_tier="T2")


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(r, ensure_ascii=False, sort_keys=True) for r in rows) + ("\n" if rows else ""), encoding="utf-8")


def merge(existing: list[dict[str, Any]], new: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id = {r.get("asset_id"): r for r in existing if r.get("asset_id")}
    for r in new:
        by_id[r["asset_id"]] = {k: v for k, v in r.items() if v is not None}
    return sorted(by_id.values(), key=lambda r: r.get("asset_id", ""))


def resolve_path(p: Path) -> Path | None:
    """Resolve common ChatGPT/container paths to local operator locations."""
    expanded = p.expanduser()
    candidates = [
        expanded,
        Path.cwd() / p.name,
        REPO / p.name,
        Path.home() / "Downloads" / p.name,
        Path.home() / "Documents" / "Data" / "Energy_Sector" / p.name,
        Path("/mnt/data") / p.name,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _safe_extract_zip(path: Path, dest: Path) -> list[Path]:
    extracted: list[Path] = []
    root = dest.resolve()
    with zipfile.ZipFile(path) as zf:
        for member in zf.infolist():
            name = member.filename
            if member.is_dir() or name.startswith("__MACOSX/") or Path(name).name.startswith("._"):
                continue
            target = (dest / name).resolve()
            if not str(target).startswith(str(root)):
                raise ValueError(f"unsafe zip member path: {name}")
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(target)
    return extracted


def _safe_extract_tar(path: Path, dest: Path) -> list[Path]:
    extracted: list[Path] = []
    root = dest.resolve()
    with tarfile.open(path) as tf:
        for member in tf.getmembers():
            if not member.isfile():
                continue
            name = member.name
            if name.startswith("__MACOSX/") or Path(name).name.startswith("._"):
                continue
            target = (dest / name).resolve()
            if not str(target).startswith(str(root)):
                raise ValueError(f"unsafe tar member path: {name}")
            target.parent.mkdir(parents=True, exist_ok=True)
            src = tf.extractfile(member)
            if src is None:
                continue
            with src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(target)
    return extracted


def _safe_extract_7z(path: Path, dest: Path) -> list[Path]:
    before = {x.resolve() for x in dest.rglob("*") if x.is_file()}
    dest.mkdir(parents=True, exist_ok=True)
    try:
        import py7zr  # type: ignore
        with py7zr.SevenZipFile(path) as zf:
            zf.extractall(dest)
    except Exception:
        # Operator fallback: if py7zr is not installed but 7z/7zz exists, use it.
        import subprocess
        exe = shutil.which("7zz") or shutil.which("7z")
        if exe is None:
            return []
        subprocess.run([exe, "x", str(path), f"-o{dest}", "-y"], check=True, stdout=subprocess.DEVNULL)
    extracted = []
    root = dest.resolve()
    for target in dest.rglob("*"):
        if not target.is_file() or target.name.startswith("._") or "__MACOSX" in target.parts:
            continue
        resolved = target.resolve()
        if not str(resolved).startswith(str(root)):
            raise ValueError(f"unsafe 7z member path: {target}")
        if resolved not in before:
            extracted.append(target)
    return extracted


def _discover_file(p: Path, extract_root: Path, report: list[dict[str, Any]], requested: Path | None = None) -> list[Path]:
    req = requested or p
    suffixes = ''.join(p.suffixes).lower()
    if p.suffix.lower() == ".zip":
        target = extract_root / p.stem
        files = _safe_extract_zip(p, target)
        expanded: list[Path] = []
        for f in files:
            if f.suffix.lower() == ".zip" or ''.join(f.suffixes).lower().endswith((".tar.gz", ".tgz")) or f.suffix.lower() == ".7z":
                expanded.extend(_discover_file(f, target / f.stem, report, req))
            elif not f.name.startswith("._"):
                expanded.append(f)
        report.append({"requested": str(req), "resolved": str(p), "status": "zip", "files": len(expanded), "extracted_to": str(target)})
        return expanded
    if suffixes.endswith(".tar.gz") or suffixes.endswith(".tgz") or p.suffix.lower() == ".tar":
        target = extract_root / p.name.replace(".tar.gz", "").replace(".tgz", "").replace(".tar", "")
        files = _safe_extract_tar(p, target)
        expanded: list[Path] = []
        for f in files:
            if f.suffix.lower() == ".zip" or ''.join(f.suffixes).lower().endswith((".tar.gz", ".tgz")) or f.suffix.lower() == ".7z":
                expanded.extend(_discover_file(f, target / f.stem, report, req))
            elif not f.name.startswith("._"):
                expanded.append(f)
        report.append({"requested": str(req), "resolved": str(p), "status": "tar", "files": len(expanded), "extracted_to": str(target)})
        return expanded
    if p.suffix.lower() == ".7z":
        target = extract_root / p.stem
        files = _safe_extract_7z(p, target)
        report.append({"requested": str(req), "resolved": str(p), "status": "7z", "files": len(files), "extracted_to": str(target), "py7zr_available": bool(files)})
        return files
    report.append({"requested": str(req), "resolved": str(p), "status": "file", "files": 1})
    return [p]


def discover(paths: list[Path], extract_root: Path) -> tuple[list[Path], list[dict[str, Any]]]:
    out: list[Path] = []
    report: list[dict[str, Any]] = []
    for requested in paths:
        p = resolve_path(requested)
        if p is None:
            report.append({"requested": str(requested), "status": "missing", "files": 0})
            continue
        if p.is_dir():
            files = [x for x in p.rglob("*") if x.is_file() and not x.name.startswith("._") and "__MACOSX" not in x.parts]
            expanded: list[Path] = []
            for f in files:
                expanded.extend(_discover_file(f, extract_root / f.stem, report, requested))
            out.extend(expanded)
            report.append({"requested": str(requested), "resolved": str(p), "status": "directory", "files": len(files), "expanded_files": len(expanded)})
        else:
            out.extend(_discover_file(p, extract_root / p.stem, report, requested))
    # Deduplicate exact paths while preserving order.
    seen: set[str] = set()
    deduped: list[Path] = []
    for item in out:
        key = str(item.resolve())
        if key not in seen:
            seen.add(key)
            deduped.append(item)
    return deduped, report

def csv_spec_for(path: Path) -> tuple[str, str, str, str] | None:
    return CSV_SPECS.get(path.name)


def inspect_file(path: Path) -> dict[str, Any]:
    item: dict[str, Any] = {"path": str(path), "name": path.name, "suffix": ''.join(path.suffixes).lower() or path.suffix.lower(), "recognized": False}
    try:
        if path.name.startswith("._") or "__MACOSX" in path.parts:
            item["skipped_reason"] = "macos_metadata"
        elif path.suffix.lower() == ".gpkg":
            layers = gpkg_layers(path)
            item["recognized_layers"] = [layer for layer in layers if _layer_spec(path, layer) is not None]
            item["capped_layers"] = {layer: MAX_IMPORT_ROWS_BY_LAYER[layer] for layer in layers if layer in MAX_IMPORT_ROWS_BY_LAYER}
            item["unmapped_layers"] = [layer for layer in layers if _layer_spec(path, layer) is None]
            item["recognized"] = bool(item["recognized_layers"])
        elif csv_spec_for(path):
            item["recognized"] = True
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                item["columns"] = reader.fieldnames or []
                item["row_count"] = sum(1 for _ in reader)
        elif path.suffix.lower() == ".shp" and path.name in SHAPEFILE_SPECS:
            item["recognized"] = True
            item["recognized_layers"] = [path.stem]
        elif path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(path, read_only=True, data_only=True)
                item["recognized_sheets"] = [s for s in wb.sheetnames if s in XLSX_SHEET_SPECS]
                item["unmapped_sheets"] = [s for s in wb.sheetnames if s not in XLSX_SHEET_SPECS]
                item["recognized"] = bool(item["recognized_sheets"])
            except Exception as exc:
                item["error"] = str(exc)
        elif path.suffix.lower() in {".geojson", ".json"} and (
            any(kw in path.name.upper() for kw in ("NID", "TRESHACIENDAS"))
        ):
            item["recognized"] = True
            doc = json.loads(path.read_text(encoding="utf-8"))
            item["feature_count"] = len(doc.get("features", []))
    except Exception as exc:
        item["error"] = str(exc)
    return item


def write_report(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    md = path.with_suffix(".md")
    lines = ["# Local Hydro Import Report", "", f"Imported assets: **{payload['imported_assets']}**", ""]
    lines.append("## Recognized files")
    for item in payload["files"]:
        if item.get("recognized"):
            lines.append(f"- `{item['name']}`")
            if item.get("recognized_layers"):
                lines.append(f"  - layers: {', '.join(item['recognized_layers'])}")
            if item.get("row_count") is not None:
                lines.append(f"  - rows: {item['row_count']}")
            if item.get("feature_count") is not None:
                lines.append(f"  - features: {item['feature_count']}")
            if item.get("recognized_sheets"):
                lines.append(f"  - sheets: {', '.join(item['recognized_sheets'])}")
    lines.append("")
    lines.append("## Missing / skipped inputs")
    for item in payload["inputs"]:
        if item.get("status") == "missing":
            lines.append(f"- missing: `{item['requested']}`")
        elif item.get("status") in {"zip", "tar", "7z"} and item.get("files", 0) == 0:
            lines.append(f"- skipped archive: `{item.get('resolved', item.get('requested'))}`")
    lines.append("")
    lines.append("## Unmapped files/layers")
    for item in payload["files"]:
        if not item.get("recognized") and not item.get("skipped_reason"):
            lines.append(f"- `{item['name']}`")
        if item.get("unmapped_layers"):
            lines.append(f"- `{item['name']}` unmapped layers: {', '.join(item['unmapped_layers'][:20])}")
        if item.get("unmapped_sheets"):
            lines.append(f"- `{item['name']}` unmapped sheets: {', '.join(item['unmapped_sheets'][:20])}")
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", action="append", type=Path, required=True, help="file, zip archive, or directory to scan")
    ap.add_argument("--out", type=Path, default=OUT)
    ap.add_argument("--report", type=Path, default=REPO / "reports" / "local_hydro_import_report.json")
    ap.add_argument("--extract-dir", type=Path, default=None, help="optional persistent extraction directory")
    args = ap.parse_args()

    temp_ctx = tempfile.TemporaryDirectory(prefix="ayl_hydro_import_") if args.extract_dir is None else None
    extract_root = args.extract_dir or Path(temp_ctx.name)
    extract_root.mkdir(parents=True, exist_ok=True)

    try:
        discovered, input_report = discover(args.src, extract_root)
        file_report = [inspect_file(path) for path in discovered]
        new: list[dict[str, Any]] = []
        counts: Counter[str] = Counter()
        for path in discovered:
            before = len(new)
            if path.name.startswith("._") or "__MACOSX" in path.parts:
                pass
            elif path.suffix.lower() == ".gpkg":
                new.extend(iter_gpkg(path))
            elif path.suffix.lower() == ".shp" and path.name in SHAPEFILE_SPECS:
                new.extend(iter_shapefile(path, SHAPEFILE_SPECS[path.name]))
            elif path.suffix.lower() == ".xlsx":
                new.extend(iter_xlsx(path))
            elif csv_spec_for(path):
                new.extend(iter_csv(path, csv_spec_for(path)))
            elif path.name == "NID_AUTH_MASTER.csv":
                new.extend(iter_csv(path, ("water", "dam", "point", "National Inventory of Dams")))
            elif path.suffix.lower() in {".geojson", ".json"} and "NID" in path.name.upper():
                new.extend(iter_geojson(path, "NID", "dam", "water"))
            elif path.suffix.lower() in {".geojson", ".json"} and "TRESHACIENDAS" in path.name.upper():
                new.extend(iter_geojson(path, "LOCAL", "historic_waterworks", "water"))
            # KARST_*/WIDL_*/ICG_* GeoJSON dropped: subsurface-corridor inference products
            # (spiderweb-pr's domain per federation_manifest.yaml), not water measurement/
            # infrastructure.
            added = len(new) - before
            if added:
                counts[path.name] += added
        new_unique = merge([], new)
        merged = merge(load_jsonl(args.out), new_unique)
        write_jsonl(args.out, merged)
        payload = {
            "candidate_assets": len(new),
            "imported_assets": len(new_unique),
            "merged_total": len(merged),
            "output": str(args.out),
            "inputs": input_report,
            "files": file_report,
            "import_counts_by_file": dict(sorted(counts.items())),
        }
        write_report(args.report, payload)
        print(f"imported {len(new_unique)} local hydro assets ({len(new)} candidates); total {len(merged)} -> {args.out}")
        print(f"wrote layer introspection report -> {args.report}")
        if len(new) == 0:
            print("warning: no local hydro assets imported; inspect report for missing paths or unmapped layers", file=sys.stderr)
        return 0
    finally:
        if temp_ctx is not None:
            temp_ctx.cleanup()


if __name__ == "__main__":
    raise SystemExit(main())
